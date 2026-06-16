import csv
import io
from datetime import date
from typing import Optional
from uuid import UUID

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Payment, Prospect
from app.schemas import PaginatedResponse, ProspectOut
from app.services import prospect_service, payment_service, employee_service


async def export_leads(
    db: AsyncSession,
    fmt: str,
    employee_id: Optional[UUID] = None,
    stage: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> StreamingResponse:
    data = await prospect_service.list_prospects(
        db,
        assigned_to=employee_id,
        stage=stage,
        page=1,
        page_size=10_000,
    )
    prospects = data.data

    if fmt == "xlsx":
        return _leads_xlsx(prospects)
    if fmt == "csv":
        return _leads_csv(prospects)
    if fmt == "pdf":
        return _leads_pdf(prospects)
    raise HTTPException(status_code=400, detail=f"Unknown format: {fmt}")


async def export_payments(
    db: AsyncSession,
    fmt: str,
    employee_id: Optional[UUID] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> StreamingResponse:
    data = await payment_service.list_payments(
        db,
        employee_id=employee_id,
        date_from=date_from,
        date_to=date_to,
        page=1,
        page_size=10_000,
    )
    payments = data.data

    if fmt == "xlsx":
        return _payments_xlsx(payments)
    if fmt == "csv":
        return _payments_csv(payments)
    raise HTTPException(status_code=400, detail=f"Unknown format: {fmt}")


# ── Excel helpers ─────────────────────────────────────────────────────────────
def _leads_xlsx(prospects: list[ProspectOut]) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "Prospect ID", "Name", "Email", "Phone",
        "Course", "Specialization", "Stage",
        "Deal Value (₹)", "Total Paid (₹)", "Payment %",
        "Assigned To", "Exam Attended", "Exam Certified",
        "Created At",
    ]
    header_fill = PatternFill("solid", fgColor="185FA5")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)

    for row_idx, p in enumerate(prospects, 2):
        ws.append([
            p.prospect_id,
            p.name,
            p.email or "",
            p.phone or "",
            p.course_name or "",
            p.specialization or "",
            p.stage,
            p.estimated_value,
            p.total_paid,
            p.payment_percentage,
            p.assigned_employee_name or "",
            "Yes" if p.exam_attended else "No",
            "Yes" if p.exam_certified else "No",
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leads.xlsx"},
    )


def _leads_csv(prospects: list[ProspectOut]) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Prospect ID", "Name", "Email", "Phone", "Course",
        "Stage", "Deal Value", "Total Paid", "Payment %",
        "Assigned To", "Created At",
    ])
    for p in prospects:
        writer.writerow([
            p.prospect_id, p.name, p.email or "", p.phone or "",
            p.course_name or "", p.stage, p.estimated_value,
            p.total_paid, p.payment_percentage,
            p.assigned_employee_name or "",
            p.created_at.strftime("%Y-%m-%d") if p.created_at else "",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads.csv"},
    )


def _leads_pdf(prospects: list[ProspectOut]) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Leads Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    table_data = [["ID", "Name", "Email", "Course", "Stage", "Deal Value", "Paid %", "Assigned To"]]
    for p in prospects:
        table_data.append([
            p.prospect_id,
            p.name[:25],
            (p.email or "")[:25],
            (p.course_name or "")[:20],
            p.stage,
            f"₹{p.estimated_value:,.0f}",
            f"{p.payment_percentage}%",
            (p.assigned_employee_name or "")[:20],
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#185FA5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=leads.pdf"},
    )


def _payments_xlsx(payments) -> StreamingResponse:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(["Date", "Lead", "Amount (₹)", "Type", "Notes", "Recorded By"])
    for p in payments:
        ws.append([
            p.payment_date.strftime("%Y-%m-%d") if p.payment_date else "",
            p.prospect_name or "",
            float(p.amount),
            p.payment_type,
            p.notes or "",
            p.created_by_name or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payments.xlsx"},
    )


def _payments_csv(payments) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Lead", "Amount", "Type", "Notes", "Recorded By"])
    for p in payments:
        writer.writerow([
            p.payment_date.strftime("%Y-%m-%d") if p.payment_date else "",
            p.prospect_name or "", float(p.amount),
            p.payment_type, p.notes or "", p.created_by_name or "",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=payments.csv"},
    )